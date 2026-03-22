"""Excel processor implementations."""

import base64
import json
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from filpe.models.job import StagedInput


class ExcelReadProcessor:
    """Processor: excel.read - read Excel file and return structured data."""

    name = "excel.read"

    def run(self, staged: StagedInput, options: dict[str, Any] | None) -> dict[str, Any]:
        """
        Read Excel file.
        Options:
          - sheet_names: list of sheet names to read (default: all)
          - max_rows: max rows per sheet (default: 10000)
          - header_row: 1-based row index for header (default: 1)
        """
        opts = options or {}
        sheet_names = opts.get("sheet_names")
        max_rows = opts.get("max_rows", 10000)
        header_row = opts.get("header_row", 1)

        wb = load_workbook(staged.path, read_only=True, data_only=True)
        try:
            sheets_to_read = sheet_names or wb.sheetnames
            result: dict[str, Any] = {"sheets": {}}

            for name in sheets_to_read:
                if name not in wb.sheetnames:
                    continue
                ws = wb[name]
                rows: list[list[Any]] = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= max_rows:
                        break
                    rows.append(list(row))

                headers: list[str] | None = None
                if rows and header_row and header_row <= len(rows):
                    headers = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(rows[header_row - 1])]
                    data_rows = rows[header_row:]
                else:
                    data_rows = rows

                if headers:
                    # Return as list of dicts
                    records = []
                    for row in data_rows:
                        rec = {}
                        for j, h in enumerate(headers):
                            if j < len(row):
                                rec[h] = row[j]
                        records.append(rec)
                    result["sheets"][name] = {"headers": headers, "rows": records}
                else:
                    result["sheets"][name] = {"rows": data_rows}

            return result
        finally:
            wb.close()


class ExcelWriteProcessor:
    """Processor: excel.write - write structured data to Excel file."""

    name = "excel.write"

    def run(self, staged: StagedInput, options: dict[str, Any] | None) -> dict[str, Any]:
        """
        Write Excel file from JSON input.
        Input: JSON file with structure matching excel.read output:
          {"sheets": {"Sheet1": {"headers": [...], "rows": [{...}, ...]}, ...}}
        Options:
          - output_filename: output file name (default: "output.xlsx")
          - sheet_order: list of sheet names for order (default: dict order)
        """
        opts = options or {}
        output_filename = opts.get("output_filename", "output.xlsx")
        if not output_filename.endswith(".xlsx"):
            output_filename += ".xlsx"
        sheet_order = opts.get("sheet_order")

        data = json.loads(staged.path.read_text(encoding="utf-8"))
        sheets_data = data.get("sheets", data) if isinstance(data, dict) else {"Sheet1": {"rows": data}}

        wb = Workbook()
        wb.remove(wb.active)

        names = sheet_order if sheet_order else list(sheets_data.keys())
        if not names:
            names = ["Sheet1"]
            sheets_data = {"Sheet1": {"rows": []}}

        for name in names:
            if name not in sheets_data:
                continue
            sh = sheets_data[name]
            ws = wb.create_sheet(title=name[:31])

            if isinstance(sh, dict) and "rows" in sh:
                rows = sh["rows"]
                headers = sh.get("headers")
                if headers:
                    ws.append(headers)
                    for row in rows:
                        ws.append([row.get(h) for h in headers])
                elif rows and isinstance(rows[0], dict):
                    headers = list(rows[0].keys())
                    ws.append(headers)
                    for row in rows:
                        ws.append([row.get(h) for h in headers])
                else:
                    for row in rows:
                        ws.append(row if isinstance(row, list) else list(row.values()))
            elif isinstance(sh, list):
                for row in sh:
                    ws.append(row if isinstance(row, list) else list(row.values()))

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        content_b64 = base64.b64encode(buf.read()).decode("ascii")

        return {
            "artifacts": [
                {
                    "name": output_filename,
                    "content_base64": content_b64,
                    "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                }
            ],
        }
